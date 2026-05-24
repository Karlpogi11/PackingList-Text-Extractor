import re
import shutil
from datetime import datetime
from pathlib import Path

import pytesseract
from pdf2image import convert_from_path
import openpyxl

BASE = Path(__file__).parent
INBOX     = BASE / "inbox"
PROCESSED = BASE / "processed"
ERRORS    = BASE / "errors"
PERMITS   = BASE / "permits"
XLSX      = BASE / "awb_log.xlsx"

HEADERS = ["HAWB", "InvoiceReference", "InvoiceTotalAmount", "DeliveryDate", "TotalQty", "ReceivedDate", "OriginalFilename", "DateLogged"]

# ── OCR ──────────────────────────────────────────────────────────────────────

def ocr_pdf(pdf_path: Path, dpi=300) -> str:
    pages = convert_from_path(str(pdf_path), dpi=dpi)
    return "\n".join(pytesseract.image_to_string(p) for p in pages)

# ── FIELD PARSING ─────────────────────────────────────────────────────────────

def parse_fields(text: str) -> dict:
    def find(patterns):
        for p in patterns:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                return m.group(1).strip()
        return ""

    hawb = find([
        r'HAWB[E]?[:\s#]+([0-9]{6,})',
        r'(?:Lon|AWB)[:\s#]+([0-9]{6,})',
        r'(?:oN|aN)\s*[=\w]+[=:\s;]+\s*([0-9]{6,})',  # OCR artifact for AWB field
    ])

    invoice_ref = find([
        r'\b(SG0\d{7,})\b',
        r'[Ss][Gg][Oo]?(0\d{7,})',
        r'[Ss][Cc](0\d{7,})',
    ])
    if invoice_ref and not invoice_ref.startswith("SG"):
        invoice_ref = "SG" + invoice_ref

    amount = find([
        r'COMMENT\s*\n?\s*([0-9,]+\.[0-9]{2})',
        r'(?:Invoice\s+)?Total[:\s]+([0-9,]+\.[0-9]{2})',
        r'Grand\s+Total[:\s]+([0-9,]+\.[0-9]{2})',
    ])
    if not amount:
        # sum the last price on each item line (e.g. "SVC,... 128.00 128.00" → 128.00)
        line_totals = re.findall(r'(\d+\.\d{2})\s*$', text, re.MULTILINE)
        if line_totals:
            total = sum(float(x) for x in line_totals)
            amount = f"{total:.2f}"

    delivery_date = find([
        r'Delivery\s+Date\s*[:\s]+(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
    ])

    total_qty = find([
        r'Total\s+Quantity\s+(\d+)',
        r'Total\s+Qty[:\s]+(\d+)',
    ])
    if not total_qty:
        item_lines = re.findall(r'^\S.*\d+\.\d{2}\s*$', text, re.MULTILINE)
        if item_lines:
            total_qty = str(len(item_lines))

    return {
        "HAWB": hawb,
        "InvoiceReference": invoice_ref,
        "InvoiceTotalAmount": amount,
        "DeliveryDate": delivery_date,
        "TotalQty": total_qty,
    }

# ── EXCEL ─────────────────────────────────────────────────────────────────────

def append_to_xlsx(row: dict):
    row["DateLogged"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    if XLSX.exists():
        wb = openpyxl.load_workbook(XLSX)
        ws = wb.active
        # add DateLogged header if missing (existing files)
        existing_headers = [c.value for c in ws[1]]
        if "DateLogged" not in existing_headers:
            ws.cell(row=1, column=len(existing_headers)+1, value="DateLogged")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADERS)
    ws.append([row.get(h, "") for h in HEADERS])
    wb.save(XLSX)

# ── MAIN PROCESS ──────────────────────────────────────────────────────────────

def process(pdf_path: Path):
    print(f"[→] Processing: {pdf_path.name}")
    try:
        text = ocr_pdf(pdf_path, dpi=300)
    except Exception as e:
        shutil.move(str(pdf_path), ERRORS / pdf_path.name)
        print(f"[ERROR] OCR failed ({e}) → moved to errors/")
        return

    # Permit check
    if re.search(r'\bpermit\b', text, re.IGNORECASE):
        fields = parse_fields(text)
        new_name = f"{fields['InvoiceReference']}.pdf" if fields['InvoiceReference'] else pdf_path.name
        dest = PERMITS / new_name
        shutil.move(str(pdf_path), dest)
        print(f"[PERMIT] {pdf_path.name} → permits/{new_name}")
        return

    fields = parse_fields(text)

    # Double-check: retry at higher DPI if fields are missing
    missing = [k for k in ["HAWB", "InvoiceTotalAmount", "DeliveryDate", "TotalQty"] if not fields[k]]
    if missing or not (fields['InvoiceReference'] or fields['HAWB']):
        print(f"[RETRY] Missing {missing}, re-OCR at 400 DPI...")
        try:
            text = ocr_pdf(pdf_path, dpi=400)
            fields = parse_fields(text)
        except Exception:
            pass

    fields["ReceivedDate"] = ""
    fields["OriginalFilename"] = pdf_path.name
    file_key = fields['InvoiceReference'] or (f"HAWB_{fields['HAWB']}" if fields['HAWB'] else "")

    missing = [k for k in ["HAWB", "InvoiceTotalAmount", "DeliveryDate", "TotalQty"] if not fields[k]]
    if missing or not file_key:
        # Rename error file to something identifiable
        label = fields['InvoiceReference'] or (f"HAWB_{fields['HAWB']}" if fields['HAWB'] else pdf_path.stem)
        error_name = f"{label}_MISSING_{'_'.join(missing)}.pdf"
        dest = ERRORS / error_name
        shutil.move(str(pdf_path), dest)
        print(f"[ERROR] Missing {missing} → errors/{error_name}")
        return

    new_name = f"{file_key}.pdf"
    month_folder = PROCESSED / datetime.now().strftime("%B %Y")
    month_folder.mkdir(parents=True, exist_ok=True)
    dest = month_folder / new_name
    if dest.exists():
        dest = month_folder / f"{dest.stem}_{int(datetime.now().timestamp())}.pdf"

    shutil.move(str(pdf_path), dest)
    append_to_xlsx(fields)
    print(f"[OK] {pdf_path.name} → {dest.relative_to(BASE)}")
    print(f"     HAWB={fields['HAWB']} | Ref={fields['InvoiceReference']} | Amt={fields['InvoiceTotalAmount']} | Date={fields['DeliveryDate']} | Qty={fields['TotalQty']}")


if __name__ == "__main__":
    # Process all PDFs currently in inbox
    pdfs = list(INBOX.glob("*.pdf")) + list(INBOX.glob("*.PDF"))
    if not pdfs:
        print("No PDFs in inbox/")
    for p in pdfs:
        process(p)
